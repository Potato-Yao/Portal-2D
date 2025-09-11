from openpyxl import load_workbook

wb = load_workbook('map1.xlsx', data_only=True)
ws = wb.active

cell = ws.cell(row=11, column=2)      # 就是 (1,1) 即 A1
c = cell.fill.fgColor
if c.type == 'rgb' and c.rgb:
    rgb = c.rgb[-6:].upper()
elif c.type == 'indexed' and c.indexed is not None:
    rgb = COLOR_INDEX[c.indexed][2:]   # 去掉 alpha
elif c.type == 'theme':
    # 主题色 1 就是黑色 000000
    THEME = ['FFFFFF','000000','EEECE1','1F497D','4F81BD',
             'C0504D','9BBB59','8064A2','4BACC6','F79646']
    theme_idx = c.theme
    rgb = THEME[theme_idx % len(THEME)]
else:
    rgb = None
print('A1 颜色:', f'#{rgb}' if rgb else '无色')